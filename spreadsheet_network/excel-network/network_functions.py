from openpyxl import load_workbook
import sys
import os
import re
import pandas as pd
import numpy as np
from tqdm import tqdm
from string import punctuation
import warnings
from pathlib import Path
from distutils.dir_util import copy_tree

def load_data(filename):
    with warnings.catch_warnings(): # turn off warning about not seeing lines and circles
        warnings.simplefilter("ignore")
        wb = load_workbook(filename = filename)
    return wb

def calculate_linked_sheets(sheet):
    cell_values_list = []
    for row in range(1, sheet.max_row):
        for column in range(1, sheet.max_column):
            this_cell = sheet.cell(row=row,column=column).value
            if this_cell is not None:
                # print(f"Cell: {column} \t Value: {this_cell}")
                cell_values_list.append(this_cell)    

    formulas_list = [cell for cell in cell_values_list if isinstance(cell, str) and cell.startswith("=")]
    sheets_linked_list = [re.findall("'.*?'", formula) for formula in formulas_list] # This is a list of lists (multiple sheets linked to per cell)
    sheets_linked = set([item for sublist in sheets_linked_list for item in sublist])
    return sheets_linked

def get_linked_sheets_dict(wb):
    print("Getting linked sheets. This may take several minutes with large workbooks...")
    linked_sheets_dict = {}
    for sheetname in tqdm(wb.sheetnames):
        sheet = wb[sheetname]
        linked_sheets_dict[sheetname] = calculate_linked_sheets(sheet)
    return linked_sheets_dict

def get_sheet_colors_dict(wb):
    print("Getting sheet colors")
    sheet_colors_dict = {}
    for sheetname in tqdm(wb.sheetnames):
        sheet = wb[sheetname]
        color_type = sheet.sheet_properties.tabColor.type
        sheet_color = getattr(sheet.sheet_properties.tabColor, color_type)
        sheet_colors_dict[sheetname] = sheet_color
    return sheet_colors_dict


def delete_sheets_to_ignore(linked_sheets_dict, sheet_colors_dict, sheets_to_ignore):
# Format links for network
    for sheetname in sheets_to_ignore:
        del linked_sheets_dict[sheetname] # edges
        del sheet_colors_dict[sheetname] # nodes
    
    return linked_sheets_dict, sheet_colors_dict

def get_edges(linked_sheets_dict):
    edges = []
    for key, value in linked_sheets_dict.items():
        if value == set():
            next
        for each_item in value:
            clean_item = each_item.translate(str.maketrans('', '', punctuation)) # remove punctuation and single quote
            edges.append([key, clean_item])

    # Reverse edges so they are From rather than To, so arrows point the right way
    edges = [list(reversed(edge)) for edge in edges]
    
    return edges

def get_nodes(sheet_colors_dict):
    # Split sheet names (the nodes) into a list of list
    # depending on colors
    nodes = []
    sub_list = []
    unique_colors = set(sheet_colors_dict.values())
    for color_num, color in enumerate(unique_colors):
        if color_num > 0:
            nodes.append(sub_list)
        sub_list = []
        for sheet_name, sheet_color in sheet_colors_dict.items():
            if sheet_color == color:
                sub_list.append(sheet_name)
    
    return nodes

# Write constants.js
def write_constants_file(edges, nodes, outfile):
    
    with open(outfile, "w") as f:
        f.write("var spreadsheet_edges = [\n")
        for item in edges:
            f.write("%s,\n" % item)
        f.write("]\n\n")

        f.write("var node_groups = [ \n")
        for item in nodes:
            f.write("%s,\n" % item)
        f.write("]\n\n")
        
    
    
def create_output_directory(outdir):
    Path(outdir).mkdir(parents=True, exist_ok=True)
    copy_tree("jsfiles/", outdir)